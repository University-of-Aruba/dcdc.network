import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
import shutil
from datetime import datetime
import os

os.chdir(r"C:\Users\Rendell CE\Cornerstone Economics\Internal - Documents\CLIENTS\CURRENT CLIENTS\GOAruba Energy\Expertise France")

src = "TCO Analysis Aruba_Complete.xlsx"
backup = f"TCO Analysis Aruba_Complete_backup_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
shutil.copy2(src, backup)
print(f"Backup: {backup}")

wb = openpyxl.load_workbook(src)

# Styles
header_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
section_font = Font(name="Calibri", size=11, bold=True, color="1F4E79")
note_font = Font(name="Calibri", size=9, italic=True, color="666666")
source_font = Font(name="Calibri", size=9, color="2E75B6")
bold_font = Font(name="Calibri", size=10, bold=True)
white_bold = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
attn_font = Font(name="Calibri", size=10, bold=True, color="C55A11")
scen_font = Font(name="Calibri", size=10, bold=True, color="548235")

input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
highlight_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

ev_color = "2E75B6"
icev_color = "C55A11"

# ============================================================
# 1. INPUT_ASSUMPTIONS
# ============================================================
ws = wb["Input_Assumptions"]

# Import duty section
ws["B93"] = "Section J: Import duty"
ws["B93"].font = section_font
ws["B95"] = "EV import duty (%)"
ws["C95"] = 0.02
ws["C95"].number_format = "0.0%"
ws["C95"].fill = input_fill
ws["E95"] = "Preferential rate for EVs (Source: DEACI/Customs)"
ws["B96"] = "ICEV import duty (%)"
ws["C96"] = 0.36
ws["C96"].number_format = "0.0%"
ws["C96"].fill = input_fill
ws["E96"] = "Standard vehicle import duty (Source: DEACI/Customs)"

# Purchase prices with duty
ws["C51"] = "=39950*(1+C95)"
ws["C51"].fill = calc_fill
ws["E51"] = "BYD Seagull catalog (39,950) * (1 + EV duty). Duty-inclusive."
ws["C52"] = "=28000*(1+C96)"
ws["C52"].fill = calc_fill
ws["E52"] = "Toyota Yaris catalog (28,000) * (1 + ICEV duty 36%). Duty-inclusive."

# Residual values
ws["C53"] = "=ROUND(C51*0.3,0)"
ws["C53"].fill = calc_fill
ws["E53"] = "EV residual ~30% of landed price (Caribbean BYD uncertain)"
ws["C54"] = "=ROUND(C52*0.3,0)"
ws["C54"].fill = calc_fill
ws["E54"] = "ICEV residual ~30% of landed price"

# Climate adjustments zeroed
ws["C63"] = 0
ws["C63"].number_format = "0%"
ws["E63"] = "Set to 0: monitored 14.39 already reflects Aruba climate. Use 10% only with manufacturer data."
ws["C64"] = 0
ws["C64"].number_format = "0.000"
ws["E64"] = "Set to 0: monitored includes A/C. Niclas ref: 0.0082 kWh/km/degC (non-monitored only)."

# Source notes
ws["E12"] = "0% flat: DEACI 2023-25 shows -5.9% CAGR. 3% retained as high scenario."
ws["E21"] = "DEACI Dec 2025 regulated pump price (full 2023-2025 series available)"
ws["E22"] = "DEACI Dec 2025 regulated pump price (full 2023-2025 series available)"

# Input classification
ws["B100"] = "INPUT CLASSIFICATION FOR USERS"
ws["B100"].font = header_font
ws["B102"] = "FIXED (change only if market conditions shift):"
ws["B102"].font = bold_font
ws["B103"] = "  Exchange rate (C17), Import duties (C95-96), Infrastructure lifespan (C85), Discount rate (C6)"
ws["B104"] = "  Vehicle specs: battery capacity, consumption, charging power, efficiency"
ws["B106"] = "REVIEW BEFORE EACH RUN:"
ws["B106"].font = attn_font
ws["B107"] = "  Fuel prices (C21-22), Electricity rates (C24-25), Annual km (C69), Lease payments (C33-34)"

# Lease payments updated per contracts (Apr 2026)
ws["C33"] = 975
ws["C33"].number_format = "#,##0"
ws["C33"].fill = input_fill
ws["E33"] = "EV monthly lease (AWG). Source: Activated Power V.B.A. contract, Feb 2025. Was 1,150."
ws["C34"] = 833
ws["C34"].number_format = "#,##0"
ws["C34"].fill = input_fill
ws["E34"] = "ICEV monthly lease (AWG). Source: Centraal Lease N.V. contract, Sep 2017. Was 925."
ws["B108"] = "  Insurance (Input_Vehicle C41/C71), Ownership type (C8), Charging strategy (C26)"
ws["B110"] = "SCENARIO VARIABLES (Scenarios tab):"
ws["B110"].font = scen_font
ws["B111"] = "  Fuel price range, Annual km range, Escalation rates"

print("1. Input_Assumptions done")

# ============================================================
# 2. CALC_EV
# ============================================================
ws = wb["Calc_EV"]

# Year 1 fixes
ws["F4"] = "=Input_Vehicle!C41"
ws["E4"] = '=IF(AND(Input_Assumptions!$C$8="Lease",Input_Assumptions!$C$35="Yes"),0,IF(Input_Vehicle!C32="Free Lifetime",0,(Input_Assumptions!$C$69/Input_Vehicle!C34)*Input_Vehicle!C33)+Input_Vehicle!C35+Input_Vehicle!C36+Input_Vehicle!C37)'

# Years 2-5
for row in range(5, 9):
    prev = row - 1
    ws[f"D{row}"] = f"=D{prev}*(1+Input_Assumptions!$C$13)"
    ws[f"E{row}"] = f"=E{prev}*(1+Input_Assumptions!$C$15)"
    ws[f"F{row}"] = f"=F{prev}*(1+Input_Assumptions!$C$14)"

print("2. Calc_EV done")

# ============================================================
# 3. CALC_ICEV
# ============================================================
ws = wb["Calc_ICEV"]

ws["F4"] = "=Input_Vehicle!C71"

for row in range(5, 9):
    prev = row - 1
    ws[f"D{row}"] = f"=D{prev}*(1+Input_Assumptions!$C$12)"
    ws[f"E{row}"] = f"=E{prev}*(1+Input_Assumptions!$C$15)"
    ws[f"F{row}"] = f"=F{prev}*(1+Input_Assumptions!$C$14)"

print("3. Calc_ICEV done")

# ============================================================
# 4. INPUT_VEHICLE
# ============================================================
ws = wb["Input_Vehicle"]
ws["D17"] = "DEACI monitored avg (Niclas 2025). Climate adj uses this as base."
ws["C19"] = 38.88
ws["C19"].number_format = "0.00"
ws["D19"] = "BYD Seagull battery capacity (kWh). Updated Apr 2026: 38.88 kWh per Energynautics (Niclas Rhein), based on DEACI/Elmar charging log analysis. Previous value 30.4 kWh inconsistent with observed max charge of 31.46 kWh at 90% efficiency."

# Charging power: corrected from 7.4 to 7.0 kW per Energynautics (Niclas, Apr 2026) based on actual charging logs
ws["C26"] = 7.0
ws["C26"].number_format = "0.0"
ws["D26"] = "Actual charging power (kW). Corrected from 7.4 to 7.0 per Energynautics (Niclas Rhein, Apr 2026) — vehicles charged at ~7 kW per DEACI monitoring data."

# ICEV fuel consumption: confirmed at 8.0 by Niclas (Apr 2026)
ws["C58"] = 8.0
ws["C58"].number_format = "0.0"
ws["D58"] = "ICEV fuel consumption (L/100km). Confirmed at 8.0 by Energynautics (Niclas Rhein, Apr 2026) after A/C research."

print("4. Input_Vehicle done")

# ============================================================
# 5. DOCUMENTATION
# ============================================================
ws = wb["Documentation"]
ws["A8"] = "v3.0 - Apr 2026: Formula fixes, Energynautics data, import duty, chart redesign\nv3.2 - Apr 15: ICEV consumption 8.4→8.0 (Niclas confirmed), charging efficiency double-count fixed, from-grid 14.39 used directly"

ws["A28"] = "DATA SOURCES"
ws["A28"].font = header_font
ws["A30"] = "Fuel prices: DEACI regulated pump prices (verkoopprijzen 2023-2025)"
ws["A31"] = "Electricity: Elmar (commercial 0.544, household 0.37 Afl/kWh)"
ws["A32"] = "EV consumption: DEACI fleet monitoring (Niclas 2025: 14.39 kWh/100km)"
ws["A33"] = "EV lease: Activated Power V.B.A. (48mo, AWG 975/mo, contract Feb 2025)"
ws["A34"] = "Annual km: DEACI monitoring (12,563 actual vs 20,000 contract)"
ws["A35"] = "Charger costs: Elmar EV ChargeUp (Gary Croes, Dec 2025)"
ws["A36"] = "Import duties: DEACI/Customs (EV 2%, ICEV 36%)"
ws["A37"] = "Climate: Niclas (0.0082 kWh/km/degC). Not applied to monitored base."
ws["A38"] = "EV: BYD Seagull Vital Edition 2024. Battery capacity (38.88 kWh) per Energynautics/DEACI charging log analysis (Apr 2026)"
ws["A39"] = "ICEV: Toyota Yaris Sedan 2023 (most common govt sedan)"

ws["A41"] = "KNOWN LIMITATIONS"
ws["A41"].font = section_font
ws["A42"] = "- Insurance premiums are estimates (no local quotes obtained)"
ws["A43"] = "- Maintenance provisions use international benchmarks"
ws["A44"] = "- BYD residual value in Caribbean market uncertain"
ws["A45"] = "- ICEV lease: Centraal Lease N.V. (48mo, AWG 833/mo, contract Sep 2017)"

ws["A47"] = "STAKEHOLDER NOTES"
ws["A47"].font = section_font
ws["A48"] = "- DOW are lease contractors for complete Government fleet"
ws["A52"] = "- Insurance NOT bundled in either lease (Art. 9 both contracts). Lessee pays separately."
ws["A53"] = "- Maintenance IS bundled in both leases (Art. 3.2 both contracts)."
ws["A49"] = "- Government needs centralized tendering for economies of scale (Maria)"
ws["A50"] = "- Consider smart charging/wallbox for controllability"
ws["A51"] = "- Dedicated metering question flagged"

for row in range(8, 11):
    ws[f"G{row}"] = None

print("5. Documentation done")

# ============================================================
# 6. SCENARIOS
# ============================================================
ws = wb["Scenarios"]
ws["D31"] = "=Input_Assumptions!C51"
ws["E31"] = "=D31"
ws["C31"] = "=Input_Assumptions!C95"
ws["C31"].number_format = "0%"
ws["D32"] = "=Input_Assumptions!C52"
ws["E32"] = "=D32"
ws["C32"] = "=Input_Assumptions!C96"
ws["C32"].number_format = "0%"
ws["B34"] = "Note: Base case prices now include import duty."
ws["B34"].font = note_font

print("6. Scenarios done")

# ============================================================
# 7. CHART TABS
# ============================================================
for name in ["Chart_Summary", "Chart_Cumulative", "Chart_Breakdown", "Chart_Sensitivity"]:
    if name in wb.sheetnames:
        del wb[name]

def style_header_row(ws, cols, row):
    for col in cols:
        ws[f"{col}{row}"].font = white_bold
        ws[f"{col}{row}"].fill = header_fill

# --- Chart_Summary ---
ws = wb.create_sheet("Chart_Summary")
ws["B1"] = "TCO SUMMARY: KEY METRICS"
ws["B1"].font = header_font

style_header_row(ws, "BCDEF", 3)
ws["B3"] = "Metric"
ws["C3"] = "Electric Vehicle"
ws["D3"] = "ICEV"
ws["E3"] = "Difference"
ws["F3"] = "% Savings"

metrics = [
    (4, "5-Year TCO (nominal)", "Calc_EV!C10", "Calc_ICEV!C10", "#,##0"),
    (5, "5-Year TCO (NPV)", "Calc_EV!C11", "Calc_ICEV!C11", "#,##0"),
    (6, "Cost per Kilometer", "Calc_EV!C12", "Calc_ICEV!C12", "#,##0.00"),
    (7, "Cost per Km (NPV)", "Calc_EV!C13", "Calc_ICEV!C13", "#,##0.00"),
]
for r, label, ev, icev, fmt in metrics:
    ws[f"B{r}"] = label
    ws[f"C{r}"] = f"={ev}"
    ws[f"C{r}"].number_format = fmt
    ws[f"D{r}"] = f"={icev}"
    ws[f"D{r}"].number_format = fmt
    ws[f"E{r}"] = f"=D{r}-C{r}"
    ws[f"E{r}"].number_format = fmt
    ws[f"F{r}"] = f"=IF(D{r}=0,0,E{r}/D{r})"
    ws[f"F{r}"].number_format = "0.0%"

ws["B9"] = "KEY ASSUMPTIONS"
ws["B9"].font = section_font
ws["B10"] = "Ownership:"
ws["C10"] = "=Input_Assumptions!C8"
ws["B11"] = "Period:"
ws["C11"] = "=Input_Assumptions!C5"
ws["D11"] = "years"
ws["B12"] = "Annual km:"
ws["C12"] = "=Input_Assumptions!C69"
ws["C12"].number_format = "#,##0"
ws["B13"] = "EV:"
ws["C13"] = "=Input_Vehicle!C7"
ws["B14"] = "ICEV:"
ws["C14"] = "=Input_Vehicle!C48"
ws["B15"] = "Discount rate:"
ws["C15"] = "=Input_Assumptions!C6"
ws["C15"].number_format = "0.0%"
ws["B16"] = "Gasoline:"
ws["C16"] = "=Input_Assumptions!C21"
ws["D16"] = "Afl/L"
ws["B17"] = "Electricity:"
ws["C17"] = "=Input_Assumptions!C28"
ws["C17"].number_format = "0.000"
ws["D17"] = "Afl/kWh"

ws["H3"] = "Vehicle"
ws["I3"] = "NPV (Afl)"
ws["H4"] = "EV"
ws["I4"] = "=Calc_EV!C11"
ws["I4"].number_format = "#,##0"
ws["H5"] = "ICEV"
ws["I5"] = "=Calc_ICEV!C11"
ws["I5"].number_format = "#,##0"

c1 = BarChart()
c1.type = "col"
c1.title = "5-Year TCO (Net Present Value)"
c1.y_axis.title = "NPV (Afl)"
c1.style = 10
c1.width = 14
c1.height = 12
cats = Reference(ws, min_col=8, min_row=4, max_row=5)
data = Reference(ws, min_col=9, min_row=3, max_row=5)
c1.add_data(data, titles_from_data=True)
c1.set_categories(cats)
c1.series[0].graphicalProperties.solidFill = "2E75B6"
c1.y_axis.numFmt = "#,##0"
c1.legend = None
ws.add_chart(c1, "B19")

ws["B36"] = "Source: TCO Analysis Aruba v3.2, Cornerstone Economics"
ws["B36"].font = source_font

print("7a. Chart_Summary done")

# --- Chart_Cumulative ---
ws = wb.create_sheet("Chart_Cumulative")
ws["B1"] = "CUMULATIVE TCO: EV vs ICEV"
ws["B1"].font = header_font

style_header_row(ws, "BCDE", 3)
ws["B3"] = "Year"
ws["C3"] = "EV Cumulative"
ws["D3"] = "ICEV Cumulative"
ws["E3"] = "Cumulative Savings"

ws["B4"] = 0
ws["C4"] = 0
ws["D4"] = 0
ws["E4"] = 0
for i in range(1, 6):
    r = 4 + i
    ws[f"B{r}"] = i
    ws[f"C{r}"] = f"=Calc_EV!I{i+3}"
    ws[f"C{r}"].number_format = "#,##0"
    ws[f"D{r}"] = f"=Calc_ICEV!I{i+3}"
    ws[f"D{r}"].number_format = "#,##0"
    ws[f"E{r}"] = f"=D{r}-C{r}"
    ws[f"E{r}"].number_format = "#,##0"

c2 = LineChart()
c2.title = "Cumulative Total Cost of Ownership"
c2.y_axis.title = "Cumulative Cost (Afl)"
c2.x_axis.title = "Year"
c2.style = 10
c2.width = 22
c2.height = 14
cats = Reference(ws, min_col=2, min_row=4, max_row=9)
c2.add_data(Reference(ws, min_col=3, min_row=3, max_row=9), titles_from_data=True)
c2.add_data(Reference(ws, min_col=4, min_row=3, max_row=9), titles_from_data=True)
c2.set_categories(cats)
c2.series[0].graphicalProperties.line.solidFill = ev_color
c2.series[0].graphicalProperties.line.width = 28000
c2.series[1].graphicalProperties.line.solidFill = icev_color
c2.series[1].graphicalProperties.line.width = 28000
c2.y_axis.numFmt = "#,##0"
c2.legend.position = "b"
ws.add_chart(c2, "B11")
ws["B28"] = "Source: TCO Analysis Aruba v3.2, Cornerstone Economics"
ws["B28"].font = source_font

print("7b. Chart_Cumulative done")

# --- Chart_Breakdown ---
ws = wb.create_sheet("Chart_Breakdown")
ws["B1"] = "ANNUAL AVERAGE COST BREAKDOWN"
ws["B1"].font = header_font

style_header_row(ws, "BCD", 3)
ws["B3"] = "Cost Category"
ws["C3"] = "EV (Afl/yr)"
ws["D3"] = "ICEV (Afl/yr)"

cost_cats = ["Ownership/Lease", "Energy/Fuel", "Maintenance", "Insurance", "Infrastructure"]
calc_cols = ["C", "D", "E", "F", "G"]
for i, (cat, col) in enumerate(zip(cost_cats, calc_cols)):
    r = 4 + i
    ws[f"B{r}"] = cat
    ws[f"C{r}"] = f"=SUM(Calc_EV!{col}4:{col}8)/5"
    ws[f"C{r}"].number_format = "#,##0"
    ws[f"D{r}"] = f"=SUM(Calc_ICEV!{col}4:{col}8)/5"
    ws[f"D{r}"].number_format = "#,##0"

ws["B9"] = "TOTAL"
ws["B9"].font = bold_font
ws["C9"] = "=SUM(C4:C8)"
ws["C9"].number_format = "#,##0"
ws["C9"].font = bold_font
ws["D9"] = "=SUM(D4:D8)"
ws["D9"].number_format = "#,##0"
ws["D9"].font = bold_font

c3 = BarChart()
c3.type = "col"
c3.title = "Annual Average Cost Composition"
c3.y_axis.title = "Annual Cost (Afl/yr)"
c3.style = 10
c3.width = 18
c3.height = 14
cats_ref = Reference(ws, min_col=2, min_row=4, max_row=8)
data = Reference(ws, min_col=3, max_col=4, min_row=3, max_row=8)
c3.add_data(data, titles_from_data=True)
c3.set_categories(cats_ref)
c3.series[0].graphicalProperties.solidFill = ev_color
c3.series[1].graphicalProperties.solidFill = icev_color
c3.y_axis.numFmt = "#,##0"
c3.legend.position = "b"
ws.add_chart(c3, "B11")
ws["B28"] = "Source: TCO Analysis Aruba v3.2, Cornerstone Economics"
ws["B28"].font = source_font

print("7c. Chart_Breakdown done")

# --- Chart_Sensitivity ---
ws = wb.create_sheet("Chart_Sensitivity")
ws["B1"] = "SENSITIVITY ANALYSIS"
ws["B1"].font = header_font

ws["B3"] = "Fuel price sensitivity"
ws["B3"].font = section_font
style_header_row(ws, "BCDE", 4)
ws["B4"] = "Gasoline (Afl/L)"
ws["C4"] = "ICEV Fuel (Afl/yr)"
ws["D4"] = "EV Elec (Afl/yr)"
ws["E4"] = "Savings"

fuel_prices = [2.00, 2.21, 2.50, 2.75, 3.00, 3.25, 3.50]
for i, fp in enumerate(fuel_prices):
    r = 5 + i
    ws[f"B{r}"] = fp
    ws[f"C{r}"] = f"=(Input_Assumptions!C69/100)*Input_Vehicle!C60*{fp}"
    ws[f"C{r}"].number_format = "#,##0"
    ws[f"D{r}"] = "=Calc_EV!D4"
    ws[f"D{r}"].number_format = "#,##0"
    ws[f"E{r}"] = f"=C{r}-D{r}"
    ws[f"E{r}"].number_format = "#,##0"
    if fp == 2.21:
        for c in "BCDE":
            ws[f"{c}{r}"].fill = highlight_fill

c4 = BarChart()
c4.type = "col"
c4.title = "Annual Energy Cost by Fuel Price"
c4.y_axis.title = "Annual Cost (Afl/yr)"
c4.x_axis.title = "Gasoline (Afl/L)"
c4.style = 10
c4.width = 18
c4.height = 12
cats_ref = Reference(ws, min_col=2, min_row=5, max_row=11)
data = Reference(ws, min_col=3, max_col=4, min_row=4, max_row=11)
c4.add_data(data, titles_from_data=True)
c4.set_categories(cats_ref)
c4.series[0].graphicalProperties.solidFill = icev_color
c4.series[1].graphicalProperties.solidFill = ev_color
c4.y_axis.numFmt = "#,##0"
c4.legend.position = "b"
ws.add_chart(c4, "B13")

ws["B30"] = "Utilization sensitivity"
ws["B30"].font = section_font
style_header_row(ws, "BCDE", 31)
ws["B31"] = "Annual Km"
ws["C31"] = "EV Energy (Afl/yr)"
ws["D31"] = "ICEV Fuel (Afl/yr)"
ws["E31"] = "Savings"

km_vals = [8000, 10000, 12563, 15000, 20000, 25000, 30000]
for i, km in enumerate(km_vals):
    r = 32 + i
    ws[f"B{r}"] = km
    ws[f"B{r}"].number_format = "#,##0"
    ws[f"C{r}"] = f"=({km}/100)*Input_Vehicle!C18*Input_Assumptions!C28"  # C18=14.39 already includes charging losses; no /C24
    ws[f"C{r}"].number_format = "#,##0"
    ws[f"D{r}"] = f"=({km}/100)*Input_Vehicle!C60*Input_Assumptions!C21"
    ws[f"D{r}"].number_format = "#,##0"
    ws[f"E{r}"] = f"=D{r}-C{r}"
    ws[f"E{r}"].number_format = "#,##0"
    if km == 12563:
        for c in "BCDE":
            ws[f"{c}{r}"].fill = highlight_fill

c5 = LineChart()
c5.title = "Annual Energy Cost by Utilization"
c5.y_axis.title = "Annual Cost (Afl/yr)"
c5.x_axis.title = "Annual Km"
c5.style = 10
c5.width = 18
c5.height = 12
cats_ref = Reference(ws, min_col=2, min_row=32, max_row=38)
data = Reference(ws, min_col=3, max_col=4, min_row=31, max_row=38)
c5.add_data(data, titles_from_data=True)
c5.set_categories(cats_ref)
c5.series[0].graphicalProperties.line.solidFill = ev_color
c5.series[0].graphicalProperties.line.width = 28000
c5.series[1].graphicalProperties.line.solidFill = icev_color
c5.series[1].graphicalProperties.line.width = 28000
c5.y_axis.numFmt = "#,##0"
c5.legend.position = "b"
ws.add_chart(c5, "B40")

ws["B57"] = "Source: TCO Analysis Aruba v3.2, Cornerstone Economics"
ws["B57"].font = source_font
ws["B58"] = "Green row = current base case"
ws["B58"].font = note_font

print("7d. Chart_Sensitivity done")

# ============================================================
# 8. RENAME OLD CHARTS
# ============================================================
if "Charts" in wb.sheetnames:
    wb["Charts"].title = "Charts_OLD"

# ============================================================
# 9. REORDER SHEETS
# ============================================================
desired = [
    "Documentation", "Input_Assumptions", "Input_Vehicle",
    "Calc_EV", "Calc_ICEV", "Comparison", "Scenarios",
    "Chart_Summary", "Chart_Cumulative", "Chart_Breakdown",
    "Chart_Sensitivity", "DataFromMonitoring", "Charts_OLD"
]
for i, name in enumerate(desired):
    if name in wb.sheetnames:
        current = wb.sheetnames.index(name)
        wb.move_sheet(name, offset=i - current)

# ============================================================
# SAVE
# ============================================================
wb.save(src)
print(f"\nSaved: {src}")
print("All fixes applied successfully.")
